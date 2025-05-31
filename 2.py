import re
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote_plus
from openpyxl import load_workbook
import time

# --- CONFIG ---
excel_path = 'Data_Capture.xlsx'
sheet_name = 'Data'
movie_col = 'B'
start_row = 3
num_movies = 6

# Output columns F–J and country/revenue K–Z
col_genre    = 'F'
col_director = 'G'
col_cast1    = 'H'
col_cast2    = 'I'
col_cast3    = 'J'
pair_cols = [
    ('K','L'), ('M','N'), ('O','P'), ('Q','R'),
    ('S','T'), ('U','V'), ('W','X'), ('Y','Z'),
]

HEADERS = {"User-Agent":"Mozilla/5.0"}

# Retry configuration
MAX_RETRIES = 3
RETRY_DELAY = 5  # seconds between retries

def get_soup_and_html(url, retries=MAX_RETRIES):
    """
    Function to get the page soup and HTML with retry mechanism.
    """
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)  # Increased timeout to 15 seconds
            print(f"GET {url} → {r.status_code}")
            r.raise_for_status()  # This will raise an exception for 4xx/5xx status codes
            return BeautifulSoup(r.text, 'html.parser'), r.text
        except requests.exceptions.Timeout:
            print(f"Timeout error while fetching {url}. Retrying ({attempt + 1}/{retries})...")
        except requests.exceptions.ConnectionError:
            print(f"Connection error while fetching {url}. Retrying ({attempt + 1}/{retries})...")
        except requests.exceptions.RequestException as e:
            print(f"Request error for {url}: {e}. Retrying ({attempt + 1}/{retries})...")
        
        time.sleep(RETRY_DELAY)  # Wait before retrying

    print(f"Failed to fetch {url} after {retries} retries.")
    return None, None

def slugify(name):
    s = re.sub(r"[^A-Za-z0-9 ]+", "", str(name))
    s = re.sub(r"\s+", " ", s).strip()
    return quote_plus(s.replace(" ", "-"))

def extract_genre(soup):
    td = soup.find('td', string=lambda t: t and 'Genre:' in t)
    return td.find_next_sibling('td').get_text(strip=True) if td else ''

def extract_director(soup):
    hdr = soup.find('h1', string=re.compile(r'Production and Technical Credits'))
    if not hdr: return ''
    tbl = hdr.find_next('table')
    for tr in tbl.find_all('tr'):
        cols = tr.find_all('td')
        if len(cols)>=3 and cols[2].get_text(strip=True).lower()=='director':
            return cols[0].get_text(strip=True)
    return ''

def extract_cast(soup):
    div = soup.find('div', class_='cast_new')
    names = [b.get_text(strip=True) for b in div.find_all('b')[:3]] if div else []
    return (names + ['', '', ''])[:3]

def extract_international_data(url):
    """
    Extracts country names and their box office revenue from the page's JavaScript.
    """
    # Send GET request to fetch the page's HTML content
    response = requests.get(url, headers=HEADERS)
    
    # Check if the page is fetched successfully
    if response.status_code != 200:
        print(f"Failed to load page. Status code: {response.status_code}")
        return []

    # Use regular expression to find the Google Charts data
    pattern = re.compile(r"google\.visualization\.arrayToDataTable\(\[(.*?)\]\);", re.DOTALL)
    match = pattern.search(response.text)
    
    if not match:
        print("No relevant data found in the page.")
        return []

    # Extract the data from the JavaScript array
    data_str = match.group(1)
    
    # Clean up the data by removing unnecessary spaces and newlines
    data_str = data_str.replace("\n", "").replace(" ", "")
    
    # Use regular expression to extract the country-revenue pairs
    country_revenue_pairs = re.findall(r"\['(.*?)',(\d+(\.\d+)?)\]", data_str)
    
    # Convert the extracted data into a list of dictionaries
    country_revenue_data = [{'Country': pair[0], 'Revenue': float(pair[1])} for pair in country_revenue_pairs]

    return country_revenue_data

# --- MAIN ---
wb = load_workbook(excel_path)
ws = wb[sheet_name]

for i in range(num_movies):
    row = start_row + i
    title = ws[f"{movie_col}{row}"].value
    if not title:
        continue

    print(f"Row {row}: {title}")
    slug = slugify(title)
    base = f"https://www.the-numbers.com/movie/{slug}"

    # Summary, Cast
    soup_sum, _ = get_soup_and_html(base + "#tab=summary")
    soup_cast, _ = get_soup_and_html(base + "#tab=cast-and-crew")
    
    if soup_sum is None or soup_cast is None:
        print(f"Skipping {title} due to failed fetch.")
        continue

    # International (using the new function to extract data)
    intl_url = f"https://www.the-numbers.com/current/cont/graphs/movie/international-iframe/{slug}"
    intl_data = extract_international_data(intl_url)

    # extract fields
    genre    = extract_genre(soup_sum)
    director = extract_director(soup_cast)
    cast1, cast2, cast3 = extract_cast(soup_cast)

    # write into Excel
    ws[f"{col_genre}{row}"].value    = genre
    ws[f"{col_director}{row}"].value = director
    ws[f"{col_cast1}{row}"].value    = cast1
    ws[f"{col_cast2}{row}"].value    = cast2
    ws[f"{col_cast3}{row}"].value    = cast3

    # Write country and revenue data into the corresponding columns
    for idx, (cc, rc) in enumerate(pair_cols):
        if idx < len(intl_data):
            country, revenue = intl_data[idx]['Country'], intl_data[idx]['Revenue']
        else:
            country, revenue = '', ''
        ws[f"{cc}{row}"].value = country
        ws[f"{rc}{row}"].value = revenue

    print(f"  → {len(intl_data)} territories scraped")

wb.save(excel_path)
print("Done – data written into F–Z, existing columns preserved.")
