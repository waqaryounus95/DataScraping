# import re
# import requests
# from bs4 import BeautifulSoup
# from urllib.parse import quote_plus
# from openpyxl import load_workbook
# import lxml.html

# # --- CONFIG ---
# excel_path = 'Data_Capture.xlsx'
# sheet_name = 'Data'
# movie_col = 'B'
# start_row = 3
# num_movies = 25

# # Output columns F–J and country/revenue K–Z
# col_genre    = 'F'
# col_director = 'G'
# col_cast1    = 'H'
# col_cast2    = 'I'
# col_cast3    = 'J'

# pair_cols = [
#     ('K','L'), ('M','N'), ('O','P'), ('Q','R'),
#     ('S','T'), ('U','V'), ('W','X'), ('Y','Z'),
# ]

# HEADERS = {"User-Agent":"Mozilla/5.0"}

# def get_soup_and_html(url):
#     r = requests.get(url, headers=HEADERS, timeout=10)
#     print(f"GET {url} → {r.status_code}")
#     r.raise_for_status()
#     return BeautifulSoup(r.text, 'html.parser'), r.text

# def slugify(name):
#     s = re.sub(r"[^A-Za-z0-9 ]+", "", str(name))
#     s = re.sub(r"\s+", " ", s).strip()
#     return quote_plus(s.replace(" ", "-"))

# def extract_genre(soup):
#     td = soup.find('td', string=lambda t: t and 'Genre:' in t)
#     return td.find_next_sibling('td').get_text(strip=True) if td else ''

# def extract_director(soup):
#     hdr = soup.find('h1', string=re.compile(r'Production and Technical Credits'))
#     if not hdr: return ''
#     tbl = hdr.find_next('table')
#     for tr in tbl.find_all('tr'):
#         cols = tr.find_all('td')
#         if len(cols)>=3 and cols[2].get_text(strip=True).lower()=='director':
#             return cols[0].get_text(strip=True)
#     return ''

# def extract_cast(soup):
#     div = soup.find('div', class_='cast_new')
#     names = [b.get_text(strip=True) for b in div.find_all('b')[:3]] if div else []
#     return (names + ['', '', ''])[:3]

# def extract_international_data(url):
#     """
#     Extracts country names and their box office revenue from the page's JavaScript.
#     """
#     # Send GET request to fetch the page's HTML content
#     response = requests.get(url, headers=HEADERS)
    
#     # Check if the page is fetched successfully
#     if response.status_code != 200:
#         print(f"Failed to load page. Status code: {response.status_code}")
#         return []

#     # Use regular expression to find the Google Charts data
#     pattern = re.compile(r"google\.visualization\.arrayToDataTable\(\[(.*?)\]\);", re.DOTALL)
#     match = pattern.search(response.text)
    
#     if not match:
#         print("No relevant data found in the page.")
#         return []

#     # Extract the data from the JavaScript array
#     data_str = match.group(1)
    
#     # Clean up the data by removing unnecessary spaces and newlines
#     data_str = data_str.replace("\n", "").replace(" ", "")
    
#     # Use regular expression to extract the country-revenue pairs
#     country_revenue_pairs = re.findall(r"\['(.*?)',(\d+(\.\d+)?)\]", data_str)
    
#     # Convert the extracted data into a list of dictionaries
#     country_revenue_data = [{'Country': pair[0], 'Revenue': float(pair[1])} for pair in country_revenue_pairs]

#     return country_revenue_data

# # --- MAIN ---
# wb = load_workbook(excel_path)
# ws = wb[sheet_name]

# for i in range(num_movies):
#     row = start_row + i
#     title = ws[f"{movie_col}{row}"].value
#     if not title:
#         continue

#     print(f"Row {row}: {title}")
#     slug = slugify(title)
#     base = f"https://www.the-numbers.com/movie/{slug}"

#     # Summary, Cast
#     soup_sum, _      = get_soup_and_html(base + "#tab=summary")
#     soup_cast, _     = get_soup_and_html(base + "#tab=cast-and-crew")
#     # International (using the new function to extract data)
#     intl_url = f"https://www.the-numbers.com/current/cont/graphs/movie/international-iframe/{slug}"
#     intl_data = extract_international_data(intl_url)

#     # extract fields
#     genre    = extract_genre(soup_sum)
#     director = extract_director(soup_cast)
#     cast1, cast2, cast3 = extract_cast(soup_cast)

#     # write into Excel
#     ws[f"{col_genre}{row}"].value    = genre
#     ws[f"{col_director}{row}"].value = director
#     ws[f"{col_cast1}{row}"].value    = cast1
#     ws[f"{col_cast2}{row}"].value    = cast2
#     ws[f"{col_cast3}{row}"].value    = cast3

#     # Write country and revenue data into the corresponding columns
#     for idx, (cc, rc) in enumerate(pair_cols):
#         if idx < len(intl_data):
#             country, revenue = intl_data[idx]['Country'], intl_data[idx]['Revenue']
#         else:
#             country, revenue = '', ''
#         ws[f"{cc}{row}"].value = country
#         ws[f"{rc}{row}"].value = revenue

#     print(f"  → {len(intl_data)} territories scraped")

# wb.save(excel_path)
# print("Done – data written into F–Z, existing columns preserved.")




import re
import time
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote_plus, urlparse
from openpyxl import load_workbook

# ------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------
excel_path   = 'Data_Capture.xlsx'
sheet_name   = 'Data'
movie_col    = 'B'
start_row    = 101
num_movies   = 15800  # change if you have more rows

# OUTPUT COLUMNS:
col_genre                = 'F'
col_director             = 'G'
col_cast1                = 'H'
col_cast2                = 'I'
col_cast3                = 'J'
col_production_countries = 'K'
col_finance              = 'L'
pair_cols = [
    ('M','N'), ('O','P'),
    ('Q','R'), ('S','T'),
    ('U','V'), ('W','X'),
    ('Y','Z'), ('AA','AB')
]

HEADERS      = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
MAX_RETRIES  = 3
RETRY_DELAY  = 5  # seconds

# ------------------------------------------------------------
# HELPERS
# ------------------------------------------------------------
def slugify(name):
    name = str(name)  # Ensure it's a string
    s = re.sub(r"[^A-Za-z0-9 \-\(\)]+", "", name)
    s = re.sub(r"\s+", " ", s).strip()
    return quote_plus(s.replace(" ", "-"))

def get_soup_and_html(url):
    for attempt in range(MAX_RETRIES):
        try:
            r = requests.get(url, headers=HEADERS, timeout=8)
            print(f"GET {url} → {r.status_code}")
            r.raise_for_status()
            return BeautifulSoup(r.text, 'html.parser'), r.text
        except requests.exceptions.RequestException as e:
            print(f"  ! Error fetching {url}: {e} (retry {attempt+1}/{MAX_RETRIES})")
            time.sleep(RETRY_DELAY)
    print(f"Failed to fetch {url} after {MAX_RETRIES} retries.")
    return None, None

def extract_canonical_slug(soup):
    tag = soup.find('meta', attrs={'property': 'og:url'})
    if not tag or not tag.get('content'):
        return None
    parsed = urlparse(tag['content'])
    return parsed.path.rstrip('/').split('/')[-1]

def fallback_search_slug(title):
    title_str = str(title)
    search_url = f"https://www.the-numbers.com/search?searchterm={quote_plus(title_str)}"
    soup_search, _ = get_soup_and_html(search_url)
    if not soup_search:
        return None

    for link in soup_search.find_all('a', href=re.compile(r"^/movie/")):
        href = link['href']
        if href.startswith("/movie/budgets"):
            continue
        return href.rstrip('/').split('/')[-1]
    return None

def extract_genre(soup):
    td = soup.find('td', string=lambda t: t and 'Genre:' in t)
    return td.find_next_sibling('td').get_text(strip=True) if td else ''

def extract_production_countries(soup):
    td = soup.find('td', string=lambda t: t and 'Production Countries:' in t)
    return td.find_next_sibling('td').get_text(";", strip=True) if td else ''

def extract_finance(soup):
    tbl = soup.find('table', id='movie_finances')
    if not tbl:
        return ''
    td = tbl.find('td', class_='data')
    return td.get_text(strip=True) if td else ''

def extract_director(soup):
    hdr = soup.find('h1', string=re.compile(r'Production and Technical Credits'))
    if not hdr:
        return ''
    tbl = hdr.find_next('table')
    for tr in tbl.find_all('tr'):
        cols = tr.find_all('td')
        if len(cols) >= 3 and cols[2].get_text(strip=True).lower() == 'director':
            return cols[0].get_text(strip=True)
    return ''

def extract_cast(soup):
    div = soup.find('div', class_='cast_new')
    if not div:
        return '', '', ''
    names = [b.get_text(strip=True) for b in div.find_all('b')[:3]]
    while len(names) < 3:
        names.append('')
    return names[0], names[1], names[2]

def extract_international_data(slug):
    url = f"https://www.the-numbers.com/current/cont/graphs/movie/international-iframe/{slug}"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        print(f"GET {url} → {r.status_code}")
        if r.status_code != 200:
            return []
    except Exception as e:
        print(f"  ! Error fetching international data: {e}")
        return []

    js = r.text
    pattern = re.compile(
        r"google\.visualization\.arrayToDataTable\(\s*\[(.*?)\]\s*\);",
        re.DOTALL
    )
    m = pattern.search(js)
    if not m:
        return []

    body = m.group(1).replace("\n", "").replace(" ", "")
    pairs = re.findall(r"\['([^']+)'\s*,\s*([\d\.]+)\s*\]", body)

    out = []
    for country, num in pairs[1:]:
        try:
            val = int(float(num))
            rev = f"${val:,}"
        except:
            rev = ''
        out.append((country, rev))
    return out

# ------------------------------------------------------------
# MAIN SCRIPT
# ------------------------------------------------------------
wb = load_workbook(excel_path)
ws = wb[sheet_name]

for i in range(num_movies):
    row = start_row + i
    title = ws[f"{movie_col}{row}"].value
    if not title:
        continue

    print(f"\nRow {row}: {title!r}")
    # Wrap each row’s work in try/except so that one failure doesn’t stop the loop:
    try:
        guessed_slug = slugify(title)
        summary_url = f"https://www.the-numbers.com/movie/{guessed_slug}#tab=summary"

        soup_sum, _ = get_soup_and_html(summary_url)
        if not soup_sum:
            print("  → Cannot fetch Summary; skipping row.")
            # Leave columns blank but continue
            continue

        correct_slug = extract_canonical_slug(soup_sum) or guessed_slug
        print(f"  → Raw canonical slug: {correct_slug}")

        if 'custom-search' in correct_slug.lower() or 'budgets' in correct_slug.lower():
            fb = fallback_search_slug(title)
            if fb:
                correct_slug = fb
                print(f"  → Fallback search slug: {correct_slug}")
            else:
                print("  → Search fallback failed; skipping row.")
                continue

        base = f"https://www.the-numbers.com/movie/{correct_slug}"
        summary_tab = base + "#tab=summary"
        cast_crew_tab = base + "#tab=cast-and-crew"

        soup_sum, _ = get_soup_and_html(summary_tab)
        soup_cast, _ = get_soup_and_html(cast_crew_tab)
        if not soup_sum or not soup_cast:
            print("  → Cannot re-fetch tabs; skipping row.")
            continue

        genre = extract_genre(soup_sum)
        prod_cty = extract_production_countries(soup_sum)
        finance = extract_finance(soup_sum)

        director = extract_director(soup_cast)
        cast1, cast2, cast3 = extract_cast(soup_cast)

        intl_data = extract_international_data(correct_slug)
        if len(intl_data) == 0:
            fb2 = fallback_search_slug(title)
            if fb2 and fb2 != correct_slug:
                print(f"  → Retrying intl data with fallback2 slug: {fb2}")
                alt_data = extract_international_data(fb2)
                if len(alt_data) > 0:
                    correct_slug = fb2
                    intl_data = alt_data
                    print(f"    → Found {len(intl_data)} territories on second fallback")
                else:
                    print("    → Second fallback still returned zero territories")
            else:
                print("  → No valid second fallback, keeping zero territories")

        # Write whatever we have into the sheet:
        ws[f"{col_genre}{row}"].value = genre
        ws[f"{col_production_countries}{row}"].value = prod_cty
        ws[f"{col_finance}{row}"].value = finance
        ws[f"{col_director}{row}"].value = director
        ws[f"{col_cast1}{row}"].value = cast1
        ws[f"{col_cast2}{row}"].value = cast2
        ws[f"{col_cast3}{row}"].value = cast3

        for idx, (c_col, r_col) in enumerate(pair_cols):
            if idx < len(intl_data):
                country, revenue = intl_data[idx]
            else:
                country, revenue = '', ''
            ws[f"{c_col}{row}"].value = country
            ws[f"{r_col}{row}"].value = revenue

        print(f"  → Written: Genre, Prod Countries, Finance, Director, Cast, {len(intl_data)} territories.")

    except Exception as e:
        print(f"  !! Unexpected error on row {row}: {e}")
     

    finally:
        try:
            wb.save(excel_path)
            print(f"  → Workbook saved up through row {row}.")
        except Exception as save_err:
            print(f"  !! Failed to save workbook after row {row}: {save_err}")

print("\nAll done – columns F–AB updated (as far as possible), original columns A–E preserved.")
