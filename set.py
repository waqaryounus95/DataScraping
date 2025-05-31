import re
import time
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote_plus, urlparse
from openpyxl import load_workbook

# ------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------
excel_path   = 'Data_Capture.xlsx'
sheet_name   = 'Data'
movie_col    = 'B'
start_row    = 3
num_movies   = 100   # change if you have more rows

# OUTPUT COLUMNS:
col_genre                = 'F'
col_director             = 'G'
col_cast1                = 'H'
col_cast2                = 'I'
col_cast3                = 'J'
col_production_countries = 'K'
col_finance              = 'L'
pair_cols = [
    ('M','N'), ('O','P'),
    ('Q','R'), ('S','T'),
    ('U','V'), ('W','X'),
    ('Y','Z'), ('AA','AB')
]

HEADERS      = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
MAX_RETRIES  = 3
RETRY_DELAY  = 5   # seconds between retries

# ------------------------------------------------------------
# HELPERS
# ------------------------------------------------------------
def slugify(name):
    """
    Produce a “guess” slug but keep hyphens and parentheses.
    e.g. "Black Panther (2018)" → "Black-Panther-(2018)"
    """
    # Remove all characters except letters, numbers, spaces, hyphens, and parentheses
    s = re.sub(r"[^A-Za-z0-9 \-\(\)]+", "", str(name))
    # Collapse multiple spaces
    s = re.sub(r"\s+", " ", s).strip()
    # Replace spaces with hyphens
    return quote_plus(s.replace(" ", "-"))

def get_soup_and_html(url):
    """
    GET the URL (with up to MAX_RETRIES on failure). Return (BeautifulSoup, raw HTML).
    If all retries fail, return (None, None).
    """
    for attempt in range(MAX_RETRIES):
        try:
            r = requests.get(url, headers=HEADERS, timeout=8)
            print(f"GET {url} → {r.status_code}")
            r.raise_for_status()
            return BeautifulSoup(r.text, 'html.parser'), r.text
        except requests.exceptions.RequestException as e:
            print(f"  ! Error fetching {url}: {e} (retry {attempt+1}/{MAX_RETRIES})")
            time.sleep(RETRY_DELAY)
    print(f"Failed to fetch {url} after {MAX_RETRIES} retries.")
    return None, None

def extract_canonical_slug(soup):
    """
    Look for <meta property="og:url" content="https://www.the-numbers.com/movie/Some-Slug">.
    Return “Some-Slug”, or None if missing.
    """
    tag = soup.find('meta', attrs={'property': 'og:url'})
    if not tag or not tag.get('content'):
        return None
    content = tag['content']  # e.g. "https://www.the-numbers.com/movie/Jurassic-World"
    parsed  = urlparse(content)
    return parsed.path.rstrip('/').split('/')[-1]

def fallback_search_slug(title):
    """
    If canonical slug is bad (or yields zero territories), go to:
      https://www.the-numbers.com/search?searchterm={quote_plus(title)}
    Pick the first <a href="/movie/..."> that is NOT '/movie/budgets',
    return its slug. If none found, return None.
    """
    search_url = f"https://www.the-numbers.com/search?searchterm={quote_plus(title)}"
    soup_search, _ = get_soup_and_html(search_url)
    if not soup_search:
        return None

    # Find the first link that starts with /movie/ and isn't /movie/budgets
    for link in soup_search.find_all('a', href=re.compile(r"^/movie/")):
        href = link['href']
        if href.startswith("/movie/budgets"):
            continue
        # Extract the slug portion
        slug = href.rstrip('/').split('/')[-1]
        return slug
    return None

def extract_genre(soup):
    td = soup.find('td', string=lambda t: t and 'Genre:' in t)
    return td.find_next_sibling('td').get_text(strip=True) if td else ''

def extract_production_countries(soup):
    td = soup.find('td', string=lambda t: t and 'Production Countries:' in t)
    if not td:
        return ''
    return td.find_next_sibling('td').get_text(";", strip=True)

def extract_finance(soup):
    tbl = soup.find('table', id='movie_finances')
    if not tbl:
        return ''
    td  = tbl.find('td', class_='data')
    return td.get_text(strip=True) if td else ''

def extract_director(soup):
    hdr = soup.find('h1', string=re.compile(r'Production and Technical Credits'))
    if not hdr:
        return ''
    tbl = hdr.find_next('table')
    for tr in tbl.find_all('tr'):
        cols = tr.find_all('td')
        if len(cols) >= 3 and cols[2].get_text(strip=True).lower() == 'director':
            return cols[0].get_text(strip=True)
    return ''

def extract_cast(soup):
    div = soup.find('div', class_='cast_new')
    if not div:
        return '', '', ''
    names = [b.get_text(strip=True) for b in div.find_all('b')[:3]]
    while len(names) < 3:
        names.append('')
    return names[0], names[1], names[2]

def extract_international_data(slug):
    """
    Hits:
      https://www.the-numbers.com/current/cont/graphs/movie/international-iframe/{slug}
    Parses the JS:
      google.visualization.arrayToDataTable([ ['Region','Box Office'], ['Japan',197062163], ... ]);
    Returns a list of (country, "$X,XXX,XXX") tuples, skipping the header row.
    """
    url = f"https://www.the-numbers.com/current/cont/graphs/movie/international-iframe/{slug}"
    r   = requests.get(url, headers=HEADERS, timeout=15)
    print(f"GET {url} → {r.status_code}")
    if r.status_code != 200:
        return []

    js = r.text
    pattern = re.compile(
        r"google\.visualization\.arrayToDataTable\(\s*\[(.*?)\]\s*\);",
        re.DOTALL
    )
    m = pattern.search(js)
    if not m:
        return []

    body  = m.group(1).replace("\n", "").replace(" ", "")
    pairs = re.findall(r"\['([^']+)'\s*,\s*([\d\.]+)\s*\]", body)

    out = []
    for country, num in pairs[1:]:  # skip header row
        try:
            val = int(float(num))
            rev = f"${val:,}"
        except:
            rev = ''
        out.append((country, rev))
    return out

# ------------------------------------------------------------
# MAIN SCRIPT
# ------------------------------------------------------------
wb = load_workbook(excel_path)
ws = wb[sheet_name]

for i in range(num_movies):
    row   = start_row + i
    title = ws[f"{movie_col}{row}"].value
    if not title:
        continue

    print(f"\nRow {row}: {title!r}")
    guessed_slug = slugify(title)
    summary_url  = f"https://www.the-numbers.com/movie/{guessed_slug}#tab=summary"

    # 1) Fetch Summary to obtain canonical slug via <meta property="og:url">
    soup_sum, _ = get_soup_and_html(summary_url)
    if not soup_sum:
        print("  → Cannot fetch Summary; skipping row.")
        continue

    correct_slug = extract_canonical_slug(soup_sum) or guessed_slug
    print(f"  → Raw canonical slug: {correct_slug}")

    # 2) If slug == "custom-search" or slug contains "search" or "budgets", fallback right away
    if 'custom-search' in correct_slug.lower() or 'budgets' in correct_slug.lower():
        fb = fallback_search_slug(title)
        if fb:
            correct_slug = fb
            print(f"  → Fallback search slug: {correct_slug}")
        else:
            print("  → Search fallback failed; skipping row.")
            continue

    # 3) Re-fetch Summary & Cast tabs using correct_slug
    base          = f"https://www.the-numbers.com/movie/{correct_slug}"
    summary_tab   = base + "#tab=summary"
    cast_crew_tab = base + "#tab=cast-and-crew"

    soup_sum, _  = get_soup_and_html(summary_tab)
    soup_cast, _ = get_soup_and_html(cast_crew_tab)
    if not soup_sum or not soup_cast:
        print("  → Cannot re-fetch tabs; skipping row.")
        continue

    # 4) Extract fields from Summary
    genre    = extract_genre(soup_sum)
    prod_cty = extract_production_countries(soup_sum)
    finance  = extract_finance(soup_sum)

    # 5) Extract from Cast & Crew
    director = extract_director(soup_cast)
    cast1, cast2, cast3 = extract_cast(soup_cast)

    # 6) Extract International data via correct_slug
    intl_data = extract_international_data(correct_slug)

    # 7) If we got zero territories, run one more fallback search
    if len(intl_data) == 0:
        fb2 = fallback_search_slug(title)
        if fb2 and fb2 != correct_slug:
            print(f"  → Retrying intl data with fallback2 slug: {fb2}")
            alt_data = extract_international_data(fb2)
            if len(alt_data) > 0:
                correct_slug = fb2
                intl_data    = alt_data
                print(f"    → Found {len(intl_data)} territories on second fallback")
            else:
                print("    → Second fallback still returned zero territories")
        else:
            print("  → No valid second fallback, keeping zero territories")

    # 8) Write all fields into Excel
    ws[f"{col_genre}{row}"]                .value = genre
    ws[f"{col_production_countries}{row}"] .value = prod_cty
    ws[f"{col_finance}{row}"]              .value = finance
    ws[f"{col_director}{row}"]             .value = director
    ws[f"{col_cast1}{row}"]                .value = cast1
    ws[f"{col_cast2}{row}"]                .value = cast2
    ws[f"{col_cast3}{row}"]                .value = cast3

    # 9) Write up to 8 (country, revenue) pairs into M–AB
    for idx, (c_col, r_col) in enumerate(pair_cols):
        if idx < len(intl_data):
            country, revenue = intl_data[idx]
        else:
            country, revenue = '', ''
        ws[f"{c_col}{row}"].value = country
        ws[f"{r_col}{row}"].value = revenue

    print(f"  → Written: Genre, Prod Countries, Finance, Director, Cast, {len(intl_data)} territories.")

# 10) Save workbook
wb.save(excel_path)
print("\nAll done – columns F–AB updated, original columns A–E preserved.")
