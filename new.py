import re
import time
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote_plus, urlparse
from openpyxl import load_workbook
from difflib import SequenceMatcher

# ------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------
excel_path   = 'Data_Capture_new.xlsx'
sheet_name   = 'Data'
movie_col    = 'B'
year_col     = 'E'
start_row    = 210
num_movies   = 4200  # adjust as needed

# OUTPUT COLUMNS:
col_genre                = 'F'
col_director             = 'G'
col_cast1                = 'H'
col_cast2                = 'I'
col_cast3                = 'J'
col_production_countries = 'K'
col_finance              = 'L'
pair_cols = [
    ('M','N'), ('O','P'),
    ('Q','R'), ('S','T'),
    ('U','V'), ('W','X'),
    ('Y','Z'), ('AA','AB')
]

HEADERS      = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
MAX_RETRIES  = 3
RETRY_DELAY  = 5  # seconds

# ------------------------------------------------------------
# HELPERS
# ------------------------------------------------------------
def slugify(name):
    name = str(name)
    s = re.sub(r"[^A-Za-z0-9 \-\(\)]+", "", name)
    s = re.sub(r"\s+", " ", s).strip()
    return quote_plus(s.replace(" ", "-"))

def get_soup_and_html(url):
    for attempt in range(MAX_RETRIES):
        try:
            r = requests.get(url, headers=HEADERS, timeout=8)
            # print(f"GET {url} → {r.status_code}")
            r.raise_for_status()
            return BeautifulSoup(r.text, 'html.parser'), r.text
        except requests.exceptions.RequestException as e:
            # print(f"  ! Error fetching {url}: {e} (retry {attempt+1}/{MAX_RETRIES})")
            time.sleep(RETRY_DELAY)
    # print(f"Failed to fetch {url} after {MAX_RETRIES} retries.")
    return None, None

def extract_canonical_slug(soup):
    tag = soup.find('meta', attrs={'property': 'og:url'})
    if not tag or not tag.get('content'):
        return None
    parsed = urlparse(tag['content'])
    return parsed.path.rstrip('/').split('/')[-1]

def fallback_search_slug(title):
    title_str = str(title)
    search_url = f"https://www.the-numbers.com/search?searchterm={quote_plus(title_str)}"
    soup_search, _ = get_soup_and_html(search_url)
    if not soup_search:
        return None

    # Try to pick a slug that seems closest by year if multiple appear:
    candidates = []
    for link in soup_search.find_all('a', href=re.compile(r"^/movie/")):
        href = link['href']
        if href.startswith("/movie/budgets"):
            continue
        slug_part = href.rstrip('/').split('/')[-1]
        candidates.append(slug_part)
    return candidates[0] if candidates else None

def extract_genre(soup):
    td = soup.find('td', string=lambda t: t and 'Genre:' in t)
    return td.find_next_sibling('td').get_text(strip=True) if td else ''

def extract_production_countries(soup):
    td = soup.find('td', string=lambda t: t and 'Production Countries:' in t)
    return td.find_next_sibling('td').get_text(";", strip=True) if td else ''

def extract_finance(soup):
    tbl = soup.find('table', id='movie_finances')
    if not tbl:
        return ''
    td = tbl.find('td', class_='data')
    return td.get_text(strip=True) if td else ''

def extract_director(soup):
    hdr = soup.find('h1', string=re.compile(r'Production and Technical Credits'))
    if not hdr:
        return ''
    tbl = hdr.find_next('table')
    for tr in tbl.find_all('tr'):
        cols = tr.find_all('td')
        if len(cols) >= 3 and cols[2].get_text(strip=True).lower() == 'director':
            return cols[0].get_text(strip=True)
    return ''

def extract_cast(soup):
    div = soup.find('div', class_='cast_new')
    if not div:
        return '', '', ''
    names = [b.get_text(strip=True) for b in div.find_all('b')[:3]]
    while len(names) < 3:
        names.append('')
    return names[0], names[1], names[2]

def extract_international_data(slug):
    url = f"https://www.the-numbers.com/current/cont/graphs/movie/international-iframe/{slug}"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        # print(f"GET {url} → {r.status_code}")
        if r.status_code != 200:
            return []
    except Exception as e:
        # print(f"  ! Error fetching international data: {e}")
        return []

    js = r.text
    pattern = re.compile(
        r"google\.visualization\.arrayToDataTable\(\s*\[(.*?)\]\s*\);",
        re.DOTALL
    )
    m = pattern.search(js)
    if not m:
        return []

    body = m.group(1).replace("\n", "").replace(" ", "")
    pairs = re.findall(r"\['([^']+)'\s*,\s*([\d\.]+)\s*\]", body)

    out = []
    for country, num in pairs[1:]:
        try:
            val = int(float(num))
            rev = f"${val:,}"
        except:
            rev = ''
        out.append((country, rev))
    return out

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def normalize_title(t):
    """
    Normalize the title string:
    - Convert to lowercase
    - Replace " ii" with " 2"
    - Strip leading/trailing whitespace
    """
    t = t.lower().strip()
    t = re.sub(r'\bii\b', '2', t)  # convert Roman "II" to "2"
    return t

def find_slug_from_custom_search(title, year):
    """
    1. Hits custom-search?searchterm=<title>
    2. Locates the <h1>Movies</h1> table
    3. Finds the row where displayed title matches & release-year matches (with fuzzy/normalized match)
    4. Returns that slug, or None.
    """
    search_term = quote_plus(str(title))
    url = f"https://www.the-numbers.com/custom-search?searchterm={search_term}"
    soup, _ = get_soup_and_html(url)
    if not soup:
        return None

    movies_h1 = soup.find("h1", string=re.compile(r"Movies", re.I))
    if not movies_h1:
        return None
    movie_table = movies_h1.find_next("table")
    if not movie_table:
        return None

    norm_target = normalize_title(str(title))  # normalized search title
    for tr in movie_table.find_all("tr"):
        if tr.find("th"):
            continue
        tds = tr.find_all("td")
        if len(tds) < 3:
            continue

        # 2nd <td>: release date text (e.g. "Nov 8, 2013")
        date_text = tds[1].get_text(strip=True)
        m = re.search(r",\s*(\d{4})$", date_text)
        m_year = m.group(1) if m else None

        # 3rd <td>: <a href="/movie/...-(2013)#tab=summary">Title</a>
        movie_link = tds[2].find("a", href=re.compile(r"^/movie/"))
        if not movie_link:
            continue
        displayed_title = movie_link.get_text(strip=True)
        norm_disp = normalize_title(displayed_title)

        # Fuzzy match on normalized titles + exact match on year
        if m_year == str(year) and similar(norm_disp, norm_target) > 0.85:
            href = movie_link["href"]
            slug = href.split("/movie/")[-1].split("#")[0]
            return slug

    return None

def parse_box_office_from_summary(summary_soup):
    """
    Given a BeautifulSoup of summary page, parse <div id="page_filling_chart"> rows
    and return [(country, total_box_office), ...].
    """
    table_div = summary_soup.find("div", id="page_filling_chart")
    if not table_div:
        return []

    out = []
    for tr in table_div.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) != 8:
            continue
        a_tag = tds[0].find("a", href=re.compile(r"#tab=box-office"))
        if not a_tag:
            continue
        country = a_tag.get_text(strip=True)
        revenue = tds[6].get_text(strip=True)
        out.append((country, revenue))
    return out

# ------------------------------------------------------------
# MAIN SCRIPT
# ------------------------------------------------------------
wb = load_workbook(excel_path)
ws = wb[sheet_name]

for i in range(num_movies):
    row   = start_row + i
    title = ws[f"{movie_col}{row}"].value
    year  = ws[f"{year_col}{row}"].value
    if not title or not year:
        continue

    try:
        # Step 1: Try year-aware custom-search to get slug
        slug = find_slug_from_custom_search(title, year)
        if slug:
            print(f"  → Found slug via custom-search: {slug}")
        else:
            # Fallback: slugify + canonical + fallback_search_slug
            guessed_slug  = slugify(title)
            summary_url   = f"https://www.the-numbers.com/movie/{guessed_slug}#tab=summary"
            soup_sum, _   = get_soup_and_html(summary_url)
            if not soup_sum:
                # print("  → Cannot fetch summary from slugify; skipping row.")
                continue

            canonical = extract_canonical_slug(soup_sum) or guessed_slug
            # print(f"  → Canonical slug: {canonical}")

            if 'custom-search' in canonical.lower() or 'budgets' in canonical.lower():
                fb = fallback_search_slug(title)
                if fb:
                    slug = fb
                    # print(f"    → fallback_search_slug: {slug}")
                else:
                    # print("    → fallback_search_slug failed; skipping row.")
                    continue
            else:
                slug = canonical

            # If slug does not contain the year, try custom-search again
            if f"({year})" not in slug:
                # print("  → Slug missing year; trying custom-search fallback…")
                slug2 = find_slug_from_custom_search(title, year)
                if slug2:
                    slug = slug2
                    # print(f"    → Corrected slug from custom-search: {slug}")
                else:
                    print(f"Could not find year-matched slug; proceeding with '{slug}'")

        # Step 2: Fetch summary & cast pages
        base_url     = f"https://www.the-numbers.com/movie/{slug}"
        summary_page = base_url + "#tab=summary"
        cast_page    = base_url + "#tab=cast-and-crew"

        soup_sum, _  = get_soup_and_html(summary_page)
        soup_cast, _ = get_soup_and_html(cast_page)
        if not soup_sum or not soup_cast:
            # print("  → Could not fetch summary/cast tabs; skipping row.")
            continue

        # Step 3: Extract static fields
        genre     = extract_genre(soup_sum)
        prod_cty  = extract_production_countries(soup_sum)
        finance   = extract_finance(soup_sum)
        director  = extract_director(soup_cast)
        cast1, cast2, cast3 = extract_cast(soup_cast)

        # Step 4: Try existing iframe approach
        intl_data = extract_international_data(slug)

        # Step 5: If iframe fails, parse from page_filling_chart
        if not intl_data:
            # print("  → No iframe data; falling back to page_filling_chart parsing.")
            fallback_pairs = parse_box_office_from_summary(soup_sum)
            if fallback_pairs:
                # print(f"    → page_filling_chart yielded {len(fallback_pairs)} countries.")
                intl_data = fallback_pairs
            else:
                print("page_filling_chart was missing or empty.")

        # Step 6: Write everything into Excel
        ws[f"{col_genre}{row}"].value                = genre
        ws[f"{col_production_countries}{row}"].value = prod_cty
        ws[f"{col_finance}{row}"].value              = finance
        ws[f"{col_director}{row}"].value             = director
        ws[f"{col_cast1}{row}"].value                = cast1
        ws[f"{col_cast2}{row}"].value                = cast2
        ws[f"{col_cast3}{row}"].value                = cast3

        # Write country/revenue pairs into M:N, O:P, …
        for idx, (c_col, r_col) in enumerate(pair_cols):
            if idx < len(intl_data):
                country, revenue = intl_data[idx]
            else:
                country, revenue = '', ''
            ws[f"{c_col}{row}"].value = country
            ws[f"{r_col}{row}"].value = revenue

        # print(f"  → Written: Genre, Prod Countries, Finance, Director, Cast, {len(intl_data)} pairs.")

    except Exception as e:
        print(f"  !! Unexpected error on row {row}: {e}")

    finally:
        try:
            wb.save(excel_path)
            # print(f"  → Workbook saved through row {row}.")
        except Exception as save_err:
            print(f"  !! Could not save after row {row}: {save_err}")

print("\nAll done – columns F–AB updated, original columns A–E preserved.")
